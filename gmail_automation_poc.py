
import openpyxl
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
import time

# Load Excel Data
excel_file = "credentials.xlsx"
wb = openpyxl.load_workbook(excel_file)
sheet = wb.active

# Configure WebDriver
service = Service("chromedriver")
options = webdriver.ChromeOptions()
options.add_argument("--start-maximized")
driver = webdriver.Chrome(service=service, options=options)

# Helper Functions
def write_result(row, result):
    sheet.cell(row=row, column=3).value = result
    wb.save(excel_file)

try:
    # Iterate through Excel rows
    for row in range(2, sheet.max_row + 1):
        email = sheet.cell(row=row, column=1).value
        password = sheet.cell(row=row, column=2).value

        # Open Gmail
        driver.get("https://mail.google.com")

        # Enter email
        email_field = WebDriverWait(driver, 10).until(
            EC.presence_of_element_located((By.ID, "identifierId"))
        )
        email_field.clear()
        email_field.send_keys(email)
        driver.find_element(By.ID, "identifierNext").click()

        # Enter password
        password_field = WebDriverWait(driver, 10).until(
            EC.presence_of_element_located((By.NAME, "password"))
        )
        password_field.clear()
        password_field.send_keys(password)
        driver.find_element(By.ID, "passwordNext").click()

        time.sleep(5)  # Wait for login process

        if "inbox" in driver.current_url:
            write_result(row, "Success")

            # Compose Email
            compose_button = WebDriverWait(driver, 10).until(
                EC.presence_of_element_located((By.CSS_SELECTOR, "div[gh='cm']"))
            )
            compose_button.click()

            to_field = WebDriverWait(driver, 10).until(
                EC.presence_of_element_located((By.NAME, "to"))
            )
            to_field.send_keys("recipient@example.com")

            subject_field = driver.find_element(By.NAME, "subjectbox")
            subject_field.send_keys("Test Subject")

            body_field = driver.find_element(By.CSS_SELECTOR, "div[aria-label='Message Body']")
            body_field.send_keys("This is a test email.")

            attach_button = driver.find_element(By.CSS_SELECTOR, "div[command='Files']")
            attach_button.click()
            time.sleep(2)  # Wait for file picker

            # Assuming file picker auto-upload implementation
            driver.execute_script("document.querySelector('input[type=file]').style.display = 'block';")
            file_input = driver.find_element(By.CSS_SELECTOR, "input[type=file]")
            file_input.send_keys("/path/to/your/resume.pdf")

            time.sleep(5)  # Wait for upload

            send_button = driver.find_element(By.CSS_SELECTOR, "div[aria-label='Send']")
            send_button.click()

            time.sleep(5)  # Wait for send completion

            # Logout
            account_button = driver.find_element(By.CSS_SELECTOR, "a[aria-label*='Account']")
            account_button.click()
            logout_button = WebDriverWait(driver, 10).until(
                EC.presence_of_element_located((By.ID, "gb_71"))
            )
            logout_button.click()

        else:
            write_result(row, "Failure")

finally:
    driver.quit()
